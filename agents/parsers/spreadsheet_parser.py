"""
Spreadsheet File Parser

Parses Excel (.xlsx, .xls) and OpenDocument Spreadsheet (.ods) files.
"""

from typing import Dict, Any, Optional, List
import logging
from pathlib import Path

from .base_parser import BaseParser, ParseResult

logger = logging.getLogger(__name__)

# Try to import spreadsheet libraries
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False

try:
    import odfpy
    ODFPY_AVAILABLE = True
except ImportError:
    ODFPY_AVAILABLE = False


class SpreadsheetParser(BaseParser):
    """Parser for spreadsheet files (Excel, ODS)."""
    
    def __init__(self):
        """Initialize spreadsheet parser."""
        super().__init__()
        
        if not OPENPYXL_AVAILABLE and not XLRD_AVAILABLE:
            logger.warning("No Excel parsing library available. Install with: pip install openpyxl or pip install xlrd")
        if not ODFPY_AVAILABLE:
            logger.warning("No ODS parsing library available. Install with: pip install odfpy")
    
    def parse(self, file_path: str, **kwargs) -> ParseResult:
        """
        Parse a spreadsheet file.
        
        Args:
            file_path: Path to spreadsheet file
            **kwargs: Additional options:
                - sheet_names: List of sheet names to parse (default: all)
                - include_headers: Include header rows (default: True)
                - include_tables: Include table structure (default: True)
                - max_rows_per_sheet: Maximum rows per sheet (optional)
        
        Returns:
            ParseResult with extracted text and table data
        """
        sheet_names = kwargs.get('sheet_names', None)
        include_headers = kwargs.get('include_headers', True)
        include_tables = kwargs.get('include_tables', True)
        max_rows_per_sheet = kwargs.get('max_rows_per_sheet', None)
        
        if not self.validate_file(file_path):
            return ParseResult(
                text='',
                error=f"File not found or invalid: {file_path}"
            )
        
        path = Path(file_path)
        ext = path.suffix.lower()
        
        if ext == '.ods':
            return self._parse_ods(file_path, sheet_names, include_headers, include_tables, max_rows_per_sheet)
        elif ext in ['.xlsx', '.xlsm']:
            return self._parse_xlsx(file_path, sheet_names, include_headers, include_tables, max_rows_per_sheet)
        elif ext == '.xls':
            return self._parse_xls(file_path, sheet_names, include_headers, include_tables, max_rows_per_sheet)
        else:
            return ParseResult(
                text='',
                error=f"Unsupported spreadsheet format: {ext}"
            )
    
    def _parse_xlsx(self, file_path: str, sheet_names: Optional[List[str]], 
                    include_headers: bool, include_tables: bool, max_rows: Optional[int]) -> ParseResult:
        """Parse .xlsx file using openpyxl."""
        if not OPENPYXL_AVAILABLE:
            return ParseResult(
                text='',
                error="openpyxl not available. Install with: pip install openpyxl"
            )
        
        try:
            import openpyxl
            
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            all_sheets_text = []
            tables = []
            metadata = {
                'filename': Path(file_path).name,
                'file_size': self.get_file_size(file_path),
                'sheet_names': workbook.sheetnames,
            }
            
            sheets_to_parse = sheet_names if sheet_names else workbook.sheetnames
            
            for sheet_name in sheets_to_parse:
                if sheet_name not in workbook.sheetnames:
                    continue
                
                sheet = workbook[sheet_name]
                sheet_text_parts = [f"=== Sheet: {sheet_name} ==="]
                sheet_rows = []
                
                for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    if max_rows and i > max_rows:
                        break
                    
                    row_values = [str(cell) if cell is not None else '' for cell in row]
                    sheet_rows.append(row_values)
                    
                    if i == 1 and include_headers:
                        sheet_text_parts.append(' | '.join(row_values))
                        sheet_text_parts.append('-' * 80)
                    elif i > 1 or not include_headers:
                        sheet_text_parts.append(' | '.join(row_values))
                
                all_sheets_text.append('\n'.join(sheet_text_parts))
                
                if include_tables and sheet_rows:
                    headers = sheet_rows[0] if include_headers and len(sheet_rows) > 0 else None
                    data_rows = sheet_rows[1:] if headers else sheet_rows
                    
                    tables.append({
                        'sheet': sheet_name,
                        'headers': headers,
                        'rows': data_rows,
                        'row_count': len(data_rows),
                        'column_count': len(headers) if headers else (len(data_rows[0]) if data_rows else 0)
                    })
            
            full_text = '\n\n'.join(all_sheets_text)
            
            return ParseResult(
                text=full_text,
                metadata=metadata,
                tables=tables
            )
        
        except Exception as e:
            logger.error(f"Error parsing XLSX file {file_path}: {e}", exc_info=True)
            return ParseResult(
                text='',
                error=f"Error parsing XLSX file: {str(e)}"
            )
    
    def _parse_xls(self, file_path: str, sheet_names: Optional[List[str]], 
                   include_headers: bool, include_tables: bool, max_rows: Optional[int]) -> ParseResult:
        """Parse .xls file using xlrd."""
        if not XLRD_AVAILABLE:
            return ParseResult(
                text='',
                error="xlrd not available. Install with: pip install xlrd"
            )
        
        try:
            import xlrd
            
            workbook = xlrd.open_workbook(file_path)
            all_sheets_text = []
            tables = []
            metadata = {
                'filename': Path(file_path).name,
                'file_size': self.get_file_size(file_path),
                'sheet_names': workbook.sheet_names(),
            }
            
            sheets_to_parse = sheet_names if sheet_names else workbook.sheet_names()
            
            for sheet_name in sheets_to_parse:
                if sheet_name not in workbook.sheet_names():
                    continue
                
                sheet = workbook.sheet_by_name(sheet_name)
                sheet_text_parts = [f"=== Sheet: {sheet_name} ==="]
                sheet_rows = []
                
                for i in range(sheet.nrows):
                    if max_rows and i >= max_rows:
                        break
                    
                    row_values = [str(sheet.cell_value(i, j)) for j in range(sheet.ncols)]
                    sheet_rows.append(row_values)
                    
                    if i == 0 and include_headers:
                        sheet_text_parts.append(' | '.join(row_values))
                        sheet_text_parts.append('-' * 80)
                    elif i > 0 or not include_headers:
                        sheet_text_parts.append(' | '.join(row_values))
                
                all_sheets_text.append('\n'.join(sheet_text_parts))
                
                if include_tables and sheet_rows:
                    headers = sheet_rows[0] if include_headers and len(sheet_rows) > 0 else None
                    data_rows = sheet_rows[1:] if headers else sheet_rows
                    
                    tables.append({
                        'sheet': sheet_name,
                        'headers': headers,
                        'rows': data_rows,
                        'row_count': len(data_rows),
                        'column_count': len(headers) if headers else (len(data_rows[0]) if data_rows else 0)
                    })
            
            full_text = '\n\n'.join(all_sheets_text)
            
            return ParseResult(
                text=full_text,
                metadata=metadata,
                tables=tables
            )
        
        except Exception as e:
            logger.error(f"Error parsing XLS file {file_path}: {e}", exc_info=True)
            return ParseResult(
                text='',
                error=f"Error parsing XLS file: {str(e)}"
            )
    
    def _parse_ods(self, file_path: str, sheet_names: Optional[List[str]], 
                   include_headers: bool, include_tables: bool, max_rows: Optional[int]) -> ParseResult:
        """Parse .ods file using odfpy."""
        if not ODFPY_AVAILABLE:
            return ParseResult(
                text='',
                error="odfpy not available. Install with: pip install odfpy"
            )
        
        try:
            from odf.opendocument import load
            from odf.table import Table, TableRow, TableCell
            from odf.text import P
            
            doc = load(file_path)
            all_sheets_text = []
            tables = []
            metadata = {
                'filename': Path(file_path).name,
                'file_size': self.get_file_size(file_path),
            }
            
            spreadsheet = doc.getElementsByType(Table)
            sheet_names_list = [sheet.getAttribute('name') for sheet in spreadsheet]
            metadata['sheet_names'] = sheet_names_list
            
            sheets_to_parse = sheet_names if sheet_names else sheet_names_list
            
            for table in spreadsheet:
                sheet_name = table.getAttribute('name')
                if sheet_name not in sheets_to_parse:
                    continue
                
                sheet_text_parts = [f"=== Sheet: {sheet_name} ==="]
                sheet_rows = []
                
                rows = table.getElementsByType(TableRow)
                for i, row in enumerate(rows):
                    if max_rows and i >= max_rows:
                        break
                    
                    cells = row.getElementsByType(TableCell)
                    row_values = []
                    for cell in cells:
                        # Get text from cell
                        text_nodes = cell.getElementsByType(P)
                        cell_text = ' '.join([str(node) for node in text_nodes]) if text_nodes else ''
                        if not cell_text:
                            cell_text = str(cell)
                        row_values.append(cell_text)
                    
                    sheet_rows.append(row_values)
                    
                    if i == 0 and include_headers:
                        sheet_text_parts.append(' | '.join(row_values))
                        sheet_text_parts.append('-' * 80)
                    elif i > 0 or not include_headers:
                        sheet_text_parts.append(' | '.join(row_values))
                
                all_sheets_text.append('\n'.join(sheet_text_parts))
                
                if include_tables and sheet_rows:
                    headers = sheet_rows[0] if include_headers and len(sheet_rows) > 0 else None
                    data_rows = sheet_rows[1:] if headers else sheet_rows
                    
                    tables.append({
                        'sheet': sheet_name,
                        'headers': headers,
                        'rows': data_rows,
                        'row_count': len(data_rows),
                        'column_count': len(headers) if headers else (len(data_rows[0]) if data_rows else 0)
                    })
            
            full_text = '\n\n'.join(all_sheets_text)
            
            return ParseResult(
                text=full_text,
                metadata=metadata,
                tables=tables
            )
        
        except Exception as e:
            logger.error(f"Error parsing ODS file {file_path}: {e}", exc_info=True)
            return ParseResult(
                text='',
                error=f"Error parsing ODS file: {str(e)}"
            )
    
    def get_supported_extensions(self) -> list:
        """Get supported file extensions."""
        return ['.xlsx', '.xls', '.xlsm', '.ods']

